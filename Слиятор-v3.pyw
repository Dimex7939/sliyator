import os
import sys
from tkinter.ttk import Progressbar, Style

import pyexcel as p
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill, Alignment
from tkinter import *
from tkinter import filedialog as fd

first_path = ""  # first file path
second_path = ""  # second file path
third_path = ""  # result file path
is_xls = False  # bool value for xls files


# functions for assigning paths №1, №2 and №3

# №1
def first_callback():
    name = fd.askopenfilename()
    global first_path
    first_path = name
    l1.configure(text=first_path)
    check()


# №2
def second_callback():
    name = fd.askopenfilename()
    global second_path
    second_path = name
    l2.configure(text=second_path)
    check()


# №3
def third_callback():
    name = fd.asksaveasfilename()
    global third_path
    third_path = name
    l3.configure(text=third_path + '.xlsx')
    check()


def converter(filepath):
    global is_xls
    spl_path = filepath.split("/")
    rm = spl_path[:-1]
    list_to_str = '/'.join([str(elem) for elem in rm])
    result_file_name = list_to_str + '/' + f'{filepath.split("/")[-1].split(".")[0]}.xlsx'
    p.save_book_as(file_name=filepath, name='Sheet1',
                   dest_file_name=result_file_name)
    is_xls = True
    return result_file_name


def delete_files(filepath):
    os.remove(filepath)
    print("File Removed!")


# read data from files and write it to result file
def four_callback():
    first_btn.configure(state=DISABLED)
    second_btn.configure(state=DISABLED)
    third_btn.configure(state=DISABLED)
    four_btn.configure(state=DISABLED)
    progress_bar.pack()
    five_btn.pack()
    root.update()

    print(first_path)
    print(second_path)
    result.configure(text="Идет загрузка файла 1...")
    if first_path.split("/")[-1].split(".")[1] == 'xls':
        wb1 = load_workbook(filename=converter(first_path))
    else:
        wb1 = load_workbook(filename=first_path)
    result.configure(text="Идет загрузка файла 2...")
    if second_path.split("/")[-1].split(".")[1] == 'xls':
        wb2 = load_workbook(filename=converter(second_path))
    else:
        wb2 = load_workbook(filename=second_path)

    sheet_name = wb1.sheetnames[0]

    sheet_ranges1 = wb1[sheet_name]
    sheet_ranges2 = wb2[sheet_name]

    wb = Workbook()
    dest_filename = f'{third_path}.xlsx'

    ws1 = wb.active
    ws1.title = "Merging"

    result.configure(text="Идет обработка файлов...")

    first_column_from_first_file = sheet_ranges1['B']
    second_column_from_first_file = sheet_ranges2['B']

    result.configure(text="Получили столбцы")

    # Наименования столбцов
    ws1['A1'] = "Наименование"
    ws1['B1'] = "Артикул"
    ws1['C1'] = "Остаток с 1 файла"
    ws1['D1'] = "Остаток со 2 файла"
    # Индекс для поочередной записи строк в файл результата
    k = 2
    # Изначальная ширина столбцов
    max_len_a = len(str(ws1['A1'].value))
    max_len_b = len(str(ws1['B1'].value))
    max_len_c = len(str(ws1['C1'].value))
    max_len_d = len(str(ws1['D1'].value))

    # Счетчик
    count = 1

    green = Color(indexed=42)
    red = Color(indexed=29)

    finded_indexes_in_2_file = []

    for i in range(1, len(first_column_from_first_file) + 1):
        article_1 = 'B' + str(i)
        name_1 = 'A' + str(i)
        find_elem = False
        for j in range(1, len(second_column_from_first_file) + 1):
            article_2 = 'B' + str(j)
            name_2 = 'A' + str(j)
            if (sheet_ranges1[name_1].value is not None or sheet_ranges1[article_1].value is not None) and \
                    sheet_ranges1[article_1].value == sheet_ranges2[article_2].value and sheet_ranges1[name_1].value == \
                    sheet_ranges2[name_2].value:
                find_elem = True
                finded_indexes_in_2_file.append(j);
                # Запись в файл результата
                ws1['A' + str(k)] = sheet_ranges1[name_1].value  # Наименование
                ws1['B' + str(k)] = sheet_ranges1[article_1].value  # Артикул
                ws1['C' + str(k)] = sheet_ranges1['C' + str(i)].value  # Остаток с 1 файла
                ws1['D' + str(k)] = sheet_ranges2['C' + str(j)].value  # Остаток с 2 файла
                # сводный столбец
                if sheet_ranges1['C' + str(i)].value:
                    if float(sheet_ranges1['C' + str(i)].value or 0) > float(sheet_ranges2['C' + str(j)].value or 0):
                        ws1['E' + str(k)] = "+"
                        filling = PatternFill(patternType='solid', fgColor=green)  # красим ячейку в зеленый
                        ws1['E' + str(k)].fill = filling
                    elif float(sheet_ranges1['C' + str(i)].value or 0) < float(sheet_ranges2['C' + str(j)].value or 0):
                        ws1['E' + str(k)] = "-"
                        filling = PatternFill(patternType='solid', fgColor=red)  # красим ячейку в красный
                        ws1['E' + str(k)].fill = filling
                ws1['E' + str(k)].alignment = Alignment(horizontal="center", vertical="center")  # центрируем текст
                # Задается новое значение ширины столбцов
                if len(str(sheet_ranges1[name_1].value)) > max_len_a:
                    max_len_a = len(str(sheet_ranges1[name_1].value))
                if len(str(sheet_ranges1[article_1].value)) > max_len_b:
                    max_len_b = len(str(sheet_ranges1[article_1].value))
                if len(str(sheet_ranges1['C' + str(i)].value)) > max_len_c:
                    max_len_c = len(str(sheet_ranges1['D' + str(i)].value))
                if len(str(sheet_ranges2['C' + str(j)].value)) > max_len_d:
                    max_len_d = len(str(sheet_ranges1['C' + str(j)].value))
                k += 1
                break

        if not find_elem:
            ws1['A' + str(k)] = sheet_ranges1[name_1].value  # Наименование
            ws1['B' + str(k)] = sheet_ranges1[article_1].value  # Артикул
            ws1['C' + str(k)] = sheet_ranges1['C' + str(i)].value  # Остаток с 1 файла
            ws1['D' + str(k)] = None
            if sheet_ranges1['C' + str(i)].value is not None:
                ws1['E' + str(k)] = "+"
                filling = PatternFill(patternType='solid', fgColor=green)  # красим ячейку в зеленый
                ws1['E' + str(k)].fill = filling
            ws1['E' + str(k)].alignment = Alignment(horizontal="center", vertical="center")  # центрируем текст
            k += 1

        if len(first_column_from_first_file) > len(second_column_from_first_file):
            min_column_from_first_file = len(second_column_from_first_file)
        else:
            min_column_from_first_file = len(first_column_from_first_file)

        if i == round(min_column_from_first_file * count / 100):
            # Олег, не работает, нужен асинхронный запрос какой-то
            # Сейчас все заебашим
            result.configure(text="Обработано " + str(i) + " строк из " + str(min_column_from_first_file))
            print("Обработано " + str(i) + " строк из " + str(min_column_from_first_file))
            progress_bar.step()
            s.configure("LabeledProgressbar", text="{0} %      ".format(count))
            root.update()
            count += 1

    result.configure(text="Обработка не нашедших столбцов со 2 файла")
    for j in range(1, len(second_column_from_first_file) + 1):
        if j not in finded_indexes_in_2_file:
            article_2 = 'B' + str(j)
            name_2 = 'A' + str(j)
            ws1['A' + str(k)] = sheet_ranges2[name_2].value  # Наименование
            ws1['B' + str(k)] = sheet_ranges2[article_2].value  # Артикул
            ws1['C' + str(k)] = None
            ws1['D' + str(k)] = sheet_ranges2['C' + str(j)].value  # Остаток с 2 файла
            if sheet_ranges2['C' + str(j)].value is not None:
                ws1['E' + str(k)] = "-"
                filling = PatternFill(patternType='solid', fgColor=red)  # красим ячейку в красный
                ws1['E' + str(k)].fill = filling
            ws1['E' + str(k)].alignment = Alignment(horizontal="center", vertical="center")  # центрируем текст
            k += 1


    # задается ширина столбцов
    ws1.column_dimensions['A'].width = max_len_a
    ws1.column_dimensions['B'].width = max_len_b
    ws1.column_dimensions['C'].width = max_len_c
    ws1.column_dimensions['D'].width = max_len_d

    if is_xls:
        delete_files(converter(first_path))
        delete_files(converter(second_path))

    wb.save(filename=dest_filename)

    result.configure(text="Выполнено")
    print("Success!")

    # close books
    wb.close()
    wb1.close()
    wb2.close()

    first_btn.configure(state=NORMAL)
    second_btn.configure(state=NORMAL)
    third_btn.configure(state=NORMAL)
    four_btn.configure(state=NORMAL)


def five_callback():
    if is_xls:
        delete_files(converter(first_path))
        delete_files(converter(second_path))
    os.execl(sys.executable, '"{}"'.format(sys.executable), *sys.argv)

def check():
    if first_path != "" and second_path != "" and third_path != "":
        four_btn.configure(state=NORMAL)
    else:
        four_btn.configure(state=DISABLED)


# main function
if __name__ == "__main__":
    root = Tk()
    root.title("Слиятор")
    root.geometry("600x400")
    # root.state('zoomed')

    first_btn = Button(text="Загрузить первый файл",  # текст кнопки
                       background="#555",  # фоновый цвет кнопки
                       foreground="#ccc",  # цвет текста
                       padx="20",  # отступ от границ до содержимого по горизонтали
                       pady="8",  # отступ от границ до содержимого по вертикали
                       font="16",  # высота шрифта
                       command=first_callback
                       )
    first_btn.pack()
    l1 = Label(text="", font="Arial 10")
    # l1.config(bd=30)
    l1.pack()

    second_btn = Button(text="Загрузить второй файл",  # текст кнопки
                        background="#555",  # фоновый цвет кнопки
                        foreground="#ccc",  # цвет текста
                        padx="20",  # отступ от границ до содержимого по горизонтали
                        pady="8",  # отступ от границ до содержимого по вертикали
                        font="16",  # высота шрифта
                        command=second_callback
                        )
    second_btn.pack()
    l2 = Label(text="", font="Arial 10")
    # l2.config(bd=30)
    l2.pack()

    third_btn = Button(text="Выбрать путь до файла с результатами",  # текст кнопки
                       background="#555",  # фоновый цвет кнопки
                       foreground="#ccc",  # цвет текста
                       padx="20",  # отступ от границ до содержимого по горизонтали
                       pady="8",  # отступ от границ до содержимого по вертикали
                       font="16",  # высота шрифта
                       command=third_callback
                       )
    third_btn.pack()
    l3 = Label(text="", font="Arial 10")
    # l3.config(bd=30)
    l3.pack()

    four_btn = Button(text="Запустить программу",  # текст кнопки
                      background="#555",  # фоновый цвет кнопки
                      foreground="#ccc",  # цвет текста
                      padx="20",  # отступ от границ до содержимого по горизонтали
                      pady="8",  # отступ от границ до содержимого по вертикали
                      font="16",  # высота шрифта
                      state=DISABLED,
                      command=four_callback
                      )
    four_btn.pack()
    result = Label(text="", foreground="#008000", font="Arial 10")
    # result.config(bd=30)
    result.pack()

    s = Style(root)
    # add the label to the progressbar style
    s.layout("LabeledProgressbar",
             [('LabeledProgressbar.trough',
               {'children': [('LabeledProgressbar.pbar',
                              {'side': 'left', 'sticky': 'ns'}),
                             ("LabeledProgressbar.label",  # label inside the bar
                              {"sticky": ""})],
                'sticky': 'nswe'})])

    progress_bar = Progressbar(root, orient="horizontal", length=300, style="LabeledProgressbar")

    five_btn = Button(text="Отмена",  # текст кнопки
                      background="#555",  # фоновый цвет кнопки
                      foreground="#ccc",  # цвет текста
                      padx="20",  # отступ от границ до содержимого по горизонтали
                      pady="8",  # отступ от границ до содержимого по вертикали
                      font="16",  # высота шрифта
                      command=five_callback
                      )

    root.mainloop()
